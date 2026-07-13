#!/usr/bin/env python3
"""Tamkoko（泰柯）月度盘点 Excel 导入脚本

输入：xlsx 文件
- 解析每个 sheet，按 period（YYYY-MM）选匹配的 sheet
- 分类列向下 filldown
- 校验：单价/数量为数字、amount = unit_price × qty 与 Excel H 列一致
- 后续 Task 8 会加 DB 写入

运行示例：
    python scripts/import_tamkoko_inventory.py \\
        tests/fixtures/tamkoko_inventory_5m.xlsx \\
        --period 2026-05
"""
from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl

_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)
from lib.importer import (
    calculate_sha256,
    get_connection,
    IngestFileManager,
    insert_batch,
)


@dataclass
class InventoryRow:
    period: str          # 'YYYY-MM'
    category: str
    sku: str
    material_name: str
    spec: str
    unit_price: float
    qty: float
    unit: str
    amount: float


_SHEET_PERIOD_RE = re.compile(r'(\d{1,2})\s*月')


def _pick_sheet(wb: openpyxl.Workbook, period: str) -> str:
    target_month = int(period.split('-')[1])
    for name in wb.sheetnames:
        m = _SHEET_PERIOD_RE.search(name)
        if m and int(m.group(1)) == target_month:
            return name
    # Fallback: 单一 sheet 文件（如 4 月盘点只有一个 'Sheet1'）时直接用
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    raise ValueError(f"No sheet matching period {period} in sheets {wb.sheetnames}")


def parse_inventory_excel(path: str, period: str) -> list[InventoryRow]:
    """
    解析 Excel → InventoryRow 列表。
    - 多 sheet 时按 period 匹配月份
    - 分类列向下 filldown
    - amount 重算（与 Excel H 列校验）
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _pick_sheet(wb, period)
    ws = wb[sheet_name]

    # 表头检测：找到含 '分类' 的行
    header_row = None
    for r in range(1, 11):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if v == '分类':
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        raise ValueError("No '分类' header found in first 10 rows")

    out: list[InventoryRow] = []
    last_category: Optional[str] = None

    for r in range(header_row + 1, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        sku = ws.cell(row=r, column=3).value
        spec = ws.cell(row=r, column=4).value
        unit_price = ws.cell(row=r, column=5).value
        qty = ws.cell(row=r, column=6).value
        unit = ws.cell(row=r, column=7).value
        amount_xls = ws.cell(row=r, column=8).value

        # 跳过完全空行
        if all(v is None or (isinstance(v, str) and not v.strip())
               for v in (cat, name, sku, unit_price, qty, amount_xls)):
            continue

        # 分类 filldown
        if cat is None or (isinstance(cat, str) and not cat.strip()):
            cat = last_category
        else:
            last_category = str(cat).strip()

        # 数字校验
        try:
            unit_price_f = float(unit_price) if unit_price is not None else None
            qty_f = float(qty) if qty is not None else None
        except (TypeError, ValueError):
            continue

        if not sku or unit_price_f is None or qty_f is None:
            continue

        # amount 重算 + 校验
        amount_calc = round(unit_price_f * qty_f, 2)
        amount_xls_f = float(amount_xls) if amount_xls is not None else amount_calc
        if abs(amount_xls_f - amount_calc) > 0.01:
            raise ValueError(
                f'Row {r} ({sku}): amount mismatch: '
                f'xlsx={amount_xls_f}, calc={amount_calc}'
            )

        out.append(InventoryRow(
            period=period,
            category=str(cat).strip(),
            sku=str(sku).strip(),
            material_name=str(name).strip() if name else '',
            spec=str(spec).strip() if spec else '',
            unit_price=unit_price_f,
            qty=qty_f,
            unit=str(unit).strip() if unit else '',
            amount=amount_calc,
        ))

    return out


def validate_rows(rows: Iterable[InventoryRow]) -> tuple[list[InventoryRow], list[tuple[InventoryRow, str]]]:
    """返回 (accepted, rejected_with_reason)。"""
    accepted: list[InventoryRow] = []
    rejected: list[tuple[InventoryRow, str]] = []
    for r in rows:
        if r.qty is None or r.unit_price is None:
            rejected.append((r, 'qty/unit_price not numeric'))
            continue
        if r.qty < 0 or r.unit_price < 0:
            rejected.append((r, 'negative qty or unit_price'))
            continue
        if not r.sku or not r.material_name:
            rejected.append((r, 'missing sku or material_name'))
            continue
        accepted.append(r)
    return accepted, rejected


def main():
    from lib.importer import setup_cli_parser
    p = setup_cli_parser('Tamkoko 月度盘点导入')
    p.add_argument('--period', required=True, help='YYYY-MM')
    p.add_argument('--store-code', default='hz_fuyang')
    args = p.parse_args()

    if not args.input:
        raise SystemExit('Usage: python import_tamkoko_inventory.py [xlsx_file] --period YYYY-MM')

    rows = parse_inventory_excel(args.input, args.period)
    accepted, rejected = validate_rows(rows)
    print(f'parsed: {len(rows)}, accepted: {len(accepted)}, rejected: {len(rejected)}')
    for r, reason in rejected:
        print(f'  REJECT {r.sku}: {reason}')

    conn = get_connection()
    try:
        summary = run_import(
            path=args.input,
            period=args.period,
            store_code=args.store_code,
            brand_code='tamkoko',
            source_type='tamkoko_inventory',
            conn=conn,
        )
        conn.commit()
        print(f'summary: {json.dumps(summary, ensure_ascii=False)}')
    except Exception as e:
        conn.rollback()
        print(f'DB import failed: {e}', file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()


# ── DB 写入 ────────────────────────────────────────────


# calculate_sha256 imported from lib.importer
# _register_source_file → IngestFileManager.create()


def _upsert_material_sku(conn, rows: list[InventoryRow]):
    """cfg.material_sku UPSERT。first_seen_period 仅在 INSERT 时写入。"""
    if not rows:
        return
    with conn.cursor() as cur:
        for r in rows:
            cur.execute("""
                INSERT INTO brand_tamkoko_cfg.material_sku
                  (sku, material_name, category, spec, unit, unit_price,
                   first_seen_period, last_seen_period, enabled)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                ON CONFLICT (sku) DO UPDATE SET
                  material_name     = EXCLUDED.material_name,
                  category          = EXCLUDED.category,
                  spec              = EXCLUDED.spec,
                  unit              = EXCLUDED.unit,
                  unit_price        = EXCLUDED.unit_price,
                  last_seen_period  = EXCLUDED.last_seen_period,
                  updated_at        = NOW()
            """, (r.sku, r.material_name, r.category, r.spec, r.unit,
                  r.unit_price, r.period, r.period))


ODS_COLUMNS = [
    "store_code", "period", "category", "sku", "material_name", "spec",
    "unit_price", "qty", "unit", "amount", "source_file_id",
]


def _write_ods(conn, rows: list[InventoryRow], source_file_id: int, store_code: str):
    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM brand_tamkoko_ods.inventory_month_end WHERE source_file_id = %s",
            (source_file_id,),
        )
    if not rows:
        return
    payload = [(
        store_code, r.period, r.category, r.sku, r.material_name, r.spec,
        r.unit_price, r.qty, r.unit, r.amount, source_file_id
    ) for r in rows]
    insert_batch(conn, "brand_tamkoko_ods.inventory_month_end", ODS_COLUMNS, payload)


def run_import(path: str, period: str, store_code: str,
               brand_code: str, source_type: str,
               conn) -> dict:
    rows = parse_inventory_excel(path, period)
    accepted, rejected = validate_rows(rows)
    file_hash = calculate_sha256(path)
    file_name = Path(path).name
    file_path = str(Path(path).resolve())
    file_size = Path(path).stat().st_size

    mgr = IngestFileManager(conn)
    existing = mgr.check(file_hash, brand_code)
    if existing:
        source_file_id = existing["id"]
        mgr.mark_pending(source_file_id, len(accepted))
    else:
        source_file_id = mgr.create(
            brand_code, store_code, source_type,
            f"{period}-01", file_name, file_path, file_hash, file_size,
        )

    _upsert_material_sku(conn, accepted)
    _write_ods(conn, accepted, source_file_id, store_code)
    mgr.mark_success(source_file_id, len(accepted))
    return {
        'ods_rows': len(accepted),
        'sku_count': len({r.sku for r in accepted}),
        'rejected': len(rejected),
        'source_file_id': source_file_id,
    }


if __name__ == '__main__':
    main()
