#!/usr/bin/env python3
"""
Yufeng 银行流水导入脚本
用途：将银行流水 Excel 文件导入到 yufeng_ods.bank_txn 表

输入：
  - 文件路径：xlsx/xls 文件
  - 目录路径：按 inputs/ 约定递归找到 yufeng/*/bank/YYYY-MM/*

功能：
  - 自动解析路径获取 brand_code/store_code/source_type/month
  - 计算 SHA-256 file_hash
  - 幂等导入：按 source_file_id 删除当次导入数据后重灌
  - 解析银行流水 Excel（表头识别、金额去逗号、时间解析、空串→NULL）
  - 更新 raw.ingest_file 状态

运行示例：
  python scripts/import_yufeng_bank_txn.py inputs/yufeng/yf_gh/bank/2025-03/银行流水_工行_250301-250731.xlsx
  python scripts/import_yufeng_bank_txn.py inputs/yufeng/yf_gh/bank/2025-03/ --dry-run
  python scripts/import_yufeng_bank_txn.py inputs/yufeng/yf_gh/bank/2025-03/ --verify
"""

import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)
from lib.importer import (
    calculate_sha256,
    get_connection,
    IngestFileManager,
    insert_batch,
    parse_path,
    setup_cli_parser,
)
from ops_logger import create_ops_logger

# Excel 列名映射（工行流水模板）
COLUMN_MAPPING = {
    "本方账号": "self_acct",
    "交易时间": "txn_time",
    "对方单位": "counterparty_name",
    "对方单位名称": "counterparty_name",
    "对方账号": "counterparty_acct",
    "借方发生额": "out_amt",
    "贷方发生额": "in_amt",
    "转入金额": "in_amt",
    "转出金额": "out_amt",
    "余额": "balance_amt",
    "摘要": "summary",
    "用途": "purpose",
    "附言": "memo",
}

SOURCE_TYPE = "bank"


def parse_path_bank(file_path: str) -> dict:
    return parse_path(file_path, SOURCE_TYPE)


def find_bank_files(input_path: str) -> list[str]:
    """
    递归查找银行流水文件
    支持：xlsx, xls
    """
    path = Path(input_path)
    files = []

    if path.is_file():
        if path.suffix.lower() in (".xlsx", ".xls"):
            files.append(str(path))
    else:
        # 递归查找 yufeng/*/bank/YYYY-MM/* 下的 Excel 文件
        for subpath in path.rglob("*"):
            if subpath.is_file() and subpath.suffix.lower() in (".xlsx", ".xls"):
                # 检查路径是否符合约定
                try:
                    parse_path_bank(str(subpath))
                    files.append(str(subpath))
                except ValueError:
                    continue

    return sorted(files)





def parse_amount(value) -> Optional[float]:
    """
    解析金额：去除逗号，转为 numeric
    空字符串/NaN/None → None
    """
    # 处理 pandas NaN 和 numpy nan
    if value is None or isinstance(value, float) and (pd.isna(value) or math.isnan(value) if hasattr(math, 'isnan') else pd.isna(value)):
        return None

    if pd.isna(value) or value == "" or value is None:
        return None

    # 处理字符串 "nan", "NaN", "NAN"
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ("", "nan", "null", "none", "-"):
            return None
        value = value.replace(",", "").strip()

    try:
        result = float(value)
        # 检查是否是有效的有限数
        if math.isnan(result) or math.isinf(result):
            return None
        return result
    except (ValueError, TypeError):
        return None


def parse_datetime(value) -> Optional[datetime]:
    """
    解析时间：支持多种格式
    空字符串/NaN → None
    """
    if pd.isna(value) or value == "" or value is None:
        return None

    # 如果已经是 datetime 类型
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.to_datetime(value).to_pydatetime()

    # 尝试解析字符串
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue

    return None


def read_bank_excel(file_path: str) -> pd.DataFrame:
    """
    读取银行流水 Excel 文件
    - 自动识别表头行
    - 转换列名
    - 清洗数据

    兼容：部分银行导出的 xlsx 的 worksheet dimension 可能错误（例如 ref="A1"），
    会导致 pandas.read_excel 只读到第一格。[典型症状：只能读到 "[HISTORYDETAIL]" ]

    为保证稳定性，这里优先用 openpyxl 逐行读取，再转成 DataFrame。
    """

    def _norm(v) -> str:
        if v is None:
            return ""
        return str(v).strip()

    # 1) 先用 openpyxl 逐行读（不依赖 dimension）
    rows: list[list[object]] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            # safety cap
            if idx >= 20000:
                break
    except Exception:
        rows = []

    header_row = None
    if rows:
        for i, row in enumerate(rows[:80]):
            vals = [_norm(v) for v in row]
            if any(v == "交易时间" for v in vals):
                header_row = i
                break

    # 2) fallback：如果 openpyxl 读取失败，再尝试 pandas（普通 xlsx/xls）
    if header_row is None:
        df_raw = pd.read_excel(file_path, header=None, nrows=10)
        for i, row in df_raw.iterrows():
            if "交易时间" in row.values:
                header_row = i
                break

        if header_row is None:
            raise ValueError(f"无法识别表头行: {file_path}")

        df = pd.read_excel(file_path, header=header_row)
    else:
        header = [_norm(v) for v in rows[header_row]]
        # trim trailing empty header cols
        last_non_empty = 0
        for j, h in enumerate(header):
            if h:
                last_non_empty = j
        header = header[: last_non_empty + 1]

        data_rows = [r[: last_non_empty + 1] for r in rows[header_row + 1 :]]
        df = pd.DataFrame(data_rows, columns=header)

    # 去除全空行
    df = df.dropna(how="all")

    # 去除标题行（如 [HISTORYDETAIL]）
    df = df[~df.iloc[:, 0].astype(str).str.contains(r"^\[.*\]$", regex=True, na=False)]

    # 重命名列
    rename_map = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str in COLUMN_MAPPING:
            rename_map[col] = COLUMN_MAPPING[col_str]

    df = df.rename(columns=rename_map)

    # 只保留需要的列
    target_cols = list(COLUMN_MAPPING.values())
    for col in target_cols:
        if col not in df.columns:
            df[col] = None

    df = df[target_cols]

    # 处理同名列（如 对方单位 和 对方单位名称 都映射到 counterparty_name）
    # 取第一个非空值合并
    dup_cols = set(c for c in df.columns if list(df.columns).count(c) > 1)
    for col in dup_cols:
        cols = [c for c in df.columns if c == col]
        if len(cols) > 1:
            df[col] = df[cols].apply(lambda row: next((v for v in row if pd.notna(v) and str(v).strip() != ''), None), axis=1)
            df = df.loc[:, ~df.columns.duplicated()]

    # 清洗数据
    df["self_acct"] = df["self_acct"].apply(lambda x: str(x).strip() if pd.notna(x) else None)
    df["txn_time"] = df["txn_time"].apply(parse_datetime)
    df["counterparty_name"] = df["counterparty_name"].apply(
        lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else None
    )
    df["counterparty_acct"] = df["counterparty_acct"].apply(
        lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else None
    )
    df["in_amt"] = df["in_amt"].apply(parse_amount)
    df["out_amt"] = df["out_amt"].apply(parse_amount)
    df["balance_amt"] = df["balance_amt"].apply(parse_amount)
    df["summary"] = df["summary"].apply(
        lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else None
    )
    df["purpose"] = df["purpose"].apply(
        lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else None
    )
    df["memo"] = df["memo"].apply(
        lambda x: str(x).strip() if pd.notna(x) and str(x).strip() else None
    )

    return df


# get_connection, IngestFileManager imported from lib.importer


def _validate_brand(brand_code: str) -> str:
    """Validate brand code format only.

    Actual existence check happens after DB connection (ops.brands).
    """
    brand_code = (brand_code or '').lower().strip()
    if not re.match(r"^[a-z][a-z0-9_]{1,31}$", brand_code):
        raise ValueError(f"非法 brand_code: {brand_code}")
    return brand_code


# Schema 命名规则：与 TypeScript API（brand-server.ts / admin/brands/route.ts）保持一致
# yufeng / bonjur → {brand}_ods / {brand}_dm；其他品牌 → brand_{brand}_ods / brand_{brand}_dm
def get_ods_schema(brand_code: str) -> str:
    if brand_code in ('yufeng', 'bonjur'):
        return f'{brand_code}_ods'
    return f'brand_{brand_code}_ods'


def get_dm_schema(brand_code: str) -> str:
    if brand_code in ('yufeng', 'bonjur'):
        return f'{brand_code}_dm'
    return f'brand_{brand_code}_dm'


def assert_brand_exists(conn, brand_code: str):
    """Ensure brand exists in ops.brands and is enabled."""
    brand_code = _validate_brand(brand_code)
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM ops.brands WHERE brand_code=%s AND enabled=true LIMIT 1",
            (brand_code,),
        )
        if cur.fetchone() is None:
            raise ValueError(f"未知或未启用的品牌 brand_code: {brand_code}")


def assert_store_exists(conn, brand_code: str, store_code: str):
    """Ensure store exists in ops.stores, is enabled, and belongs to the given brand."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM ops.stores WHERE brand_code=%s AND store_code=%s AND enabled=true LIMIT 1",
            (brand_code, store_code),
        )
        if cur.fetchone() is None:
            # Fetch valid stores for the error message
            cur.execute(
                "SELECT store_code, store_name FROM ops.stores WHERE brand_code=%s AND enabled=true ORDER BY store_code",
                (brand_code,),
            )
            valid = ", ".join(f"{r[0]}({r[1]})" for r in cur.fetchall())
            raise ValueError(
                f"store_code '{store_code}' is not a valid enabled store for brand '{brand_code}'. "
                f"Valid stores: {valid}"
            )


def delete_existing_data(brand_code: str, source_file_id: int, conn):
    """删除当次导入的旧数据（幂等核心）"""
    brand_code = _validate_brand(brand_code)
    table = f"{get_ods_schema(brand_code)}.bank_txn"
    with conn.cursor() as cur:
        cur.execute(
            f"DELETE FROM {table} WHERE source_file_id = %s",
            (source_file_id,),
        )
        deleted = cur.rowcount
        conn.commit()
        return deleted


def insert_bank_txn(brand_code: str, df: pd.DataFrame, store_code: str, source_file_id: int, conn) -> int:
    """批量插入银行流水数据"""
    brand_code = _validate_brand(brand_code)
    table = f"{get_ods_schema(brand_code)}.bank_txn"
    records = []
    for _, row in df.iterrows():
        # 安全检查：确保金额字段不是 NaN/inf，转换为 None
        def sanitize_amount(val):
            if val is None:
                return None
            if isinstance(val, float):
                if pd.isna(val) or math.isnan(val) or math.isinf(val):
                    return None
            return val

        records.append(
            (
                store_code,
                row["self_acct"],
                row["txn_time"],
                row["counterparty_name"],
                row["counterparty_acct"],
                sanitize_amount(row["in_amt"]),
                sanitize_amount(row["out_amt"]),
                sanitize_amount(row["balance_amt"]),
                row["summary"],
                row["purpose"],
                row["memo"],
                source_file_id,
            )
        )

    if not records:
        return 0

    with conn.cursor() as cur:
        execute_values(
            cur,
            f"""
            INSERT INTO {table} (
                store_code, self_acct, txn_time, counterparty_name, counterparty_acct,
                in_amt, out_amt, balance_amt, summary, purpose, memo, source_file_id
            ) VALUES %s
            """,
            records,
        )
        conn.commit()
        return len(records)



def dry_run_import(file_path: str) -> dict:
    """
    干运行模式：解析并验证数据，不写入数据库
    """
    print(f"\n=== Dry Run: {file_path} ===")

    meta = parse_path_bank(file_path)
    print(f"Brand Code: {meta['brand_code']}")
    print(f"Store Code: {meta['store_code']}")
    print(f"Source Type: {meta['source_type']}")
    print(f"Month: {meta['month']}")
    print(f"File Name: {meta['file_name']}")

    file_hash = calculate_sha256(file_path)
    file_size = os.path.getsize(file_path)
    print(f"File Hash: {file_hash}")
    print(f"File Size: {file_size} bytes")

    # 读取并解析 Excel
    df = read_bank_excel(file_path)
    print(f"\nParsed {len(df)} rows from Excel")

    # 数据质量检查
    print("\n=== Data Quality Check ===")
    print(f"Total rows: {len(df)}")
    print(f"Non-null txn_time: {df['txn_time'].notna().sum()}")
    print(f"Non-null in_amt: {df['in_amt'].notna().sum()}")
    print(f"Non-null out_amt: {df['out_amt'].notna().sum()}")
    print(f"Non-null balance_amt: {df['balance_amt'].notna().sum()}")

    print("\n=== Sample Data (first 3 rows) ===")
    print(df.head(3).to_string())

    return {
        "meta": meta,
        "file_hash": file_hash,
        "file_size": file_size,
        "row_count": len(df),
    }


def verify_import(file_path: str) -> dict:
    """
    验证模式：检查数据是否已导入及回溯关系
    """
    print(f"\n=== Verify: {file_path} ===")

    meta = parse_path_bank(file_path)

    file_hash = calculate_sha256(file_path)
    print(f"File Hash: {file_hash}")

    conn = get_connection()

    try:
        mgr = IngestFileManager(conn)
        ingest = mgr.check(file_hash)
        if not ingest:
            print("⚠️  No ingest_file record found (not imported yet)")
            return {"status": "not_imported"}

        print(f"\n=== Ingest File Record ===")
        print(f"ID: {ingest['id']}")
        print(f"Status: {ingest['status']}")
        print(f"Row Count: {ingest['row_count']}")
        print(f"Brand: {ingest['brand_code']}")
        print(f"Store: {ingest['store_code']}")
        print(f"Month: {ingest['month']}")

        # 检查 bank_txn
        brand_code = _validate_brand(meta["brand_code"])
        table = f"{get_ods_schema(brand_code)}.bank_txn"
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT COUNT(*), COUNT(txn_time), COUNT(in_amt), COUNT(out_amt)
                FROM {table}
                WHERE source_file_id = %s
                """,
                (ingest["id"],),
            )
            row = cur.fetchone()

        print(f"\n=== Bank Txn Data ===")
        print(f"Total rows: {row[0]}")
        print(f"Non-null txn_time: {row[1]}")
        print(f"Non-null in_amt: {row[2]}")
        print(f"Non-null out_amt: {row[3]}")

        if row[0] != ingest["row_count"]:
            print(f"⚠️  Warning: bank_txn rows ({row[0]}) != ingest_file.row_count ({ingest['row_count']})")

        return {
            "status": "verified",
            "ingest_id": ingest["id"],
            "ingest_status": ingest["status"],
            "bank_txn_count": row[0],
            "expected_count": ingest["row_count"],
        }

    finally:
        conn.close()


def do_import(file_path: str) -> dict:
    """
    执行导入
    """
    print(f"\n=== Importing: {file_path} ===")

    meta = parse_path_bank(file_path)

    _validate_brand(meta["brand_code"])

    ops = create_ops_logger(
        brand_code=meta["brand_code"],
        store_code=meta["store_code"],
        month=meta["month"],
        triggered_by="manual",
        note=meta["file_name"],
    )

    file_hash = calculate_sha256(file_path)
    file_size = os.path.getsize(file_path)
    print(f"File Hash: {file_hash}")

    conn = get_connection()

    assert_brand_exists(conn, meta["brand_code"])
    assert_store_exists(conn, meta["brand_code"], meta["store_code"])

    STEP_ORDER = {
        "register_file": 1,
        "delete_previous": 2,
        "load_excel": 3,
        "insert_bank_txn": 4,
        "update_ingest_status": 5,
        "refresh_snapshot": 6,
    }

    try:
        if ops:
            ops.step_start("register_file", step_order=STEP_ORDER["register_file"], detail={"file_path": file_path})

        mgr = IngestFileManager(conn)
        existing = mgr.check(file_hash)

        source_file_id = None
        if existing:
            print(f"Found existing import: id={existing['id']}, status={existing['status']}")
            source_file_id = existing["id"]

            if existing["status"] == "success":
                if ops:
                    ops.step_start("delete_previous", step_order=STEP_ORDER["delete_previous"])

                deleted = delete_existing_data(meta["brand_code"], source_file_id, conn)
                print(f"Deleted {deleted} existing rows for idempotent import")

                if ops:
                    ops.step_end("delete_previous", rows_out=deleted)
            elif existing["status"] == "pending":
                print("Warning: previous import is pending, will retry")
        else:
            source_file_id = mgr.create(
                meta["brand_code"], meta["store_code"], SOURCE_TYPE,
                meta["month_date"], meta["file_name"], meta["file_path"],
                file_hash, file_size,
            )
            print(f"Created new ingest_file: id={source_file_id}")

        if ops:
            ops.step_end("register_file", rows_out=1, detail={"source_file_id": source_file_id})

        if ops:
            ops.step_start("load_excel", step_order=STEP_ORDER["load_excel"])

        df = read_bank_excel(file_path)
        rows_parsed = len(df)
        print(f"Parsed {rows_parsed} rows from Excel")

        if ops:
            ops.step_end("load_excel", rows_out=rows_parsed)

        if ops:
            ops.step_start("insert_bank_txn", step_order=STEP_ORDER["insert_bank_txn"], detail={"rows_in": rows_parsed})

        row_count = insert_bank_txn(meta["brand_code"], df, meta["store_code"], source_file_id, conn)
        print(f"Inserted {row_count} rows into {get_ods_schema(meta['brand_code'])}.bank_txn")

        if ops:
            ops.step_end("insert_bank_txn", rows_out=row_count, rows_rejected=rows_parsed - row_count)

        if ops:
            ops.step_start("update_ingest_status", step_order=STEP_ORDER["update_ingest_status"])

        mgr.mark_success(source_file_id, row_count)
        print(f"Updated ingest_file status to success")

        if ops:
            ops.step_end("update_ingest_status", rows_out=1)

        try:
            if ops:
                ops.step_start("refresh_snapshot", step_order=STEP_ORDER["refresh_snapshot"], detail={"source_file_id": source_file_id})

            brand_code = _validate_brand(meta["brand_code"])
            with conn.cursor() as cur:
                cur.execute(f"SELECT {get_dm_schema(brand_code)}.refresh_bank_txn_classified_snapshot(%s)", (source_file_id,))
            conn.commit()

            if ops:
                ops.step_end("refresh_snapshot", rows_out=row_count)
        except Exception as e:
            msg = f"WARN: refresh snapshot skipped: {e}"
            print(msg)
            if ops:
                ops.step_end("refresh_snapshot", status="skipped", error_message=str(e)[:500])

        if ops:
            ops.finish(status="success")

        return {
            "status": "success",
            "source_file_id": source_file_id,
            "row_count": row_count,
        }

    except Exception as e:
        error_msg = str(e)
        print(f"Error: {error_msg}")

        if ops:
            ops.step_end("insert_bank_txn", status="failed", error_message=error_msg[:500])
            ops.finish(status="failed", note=error_msg[:500])

        if source_file_id:
            mgr.mark_failed(source_file_id, error_msg[:1000])

        raise

    finally:
        conn.close()
        if ops:
            ops.close()


def main():
    parser = setup_cli_parser("Yufeng 银行流水导入脚本")
    parser.add_argument(
        "input_path",
        help="文件路径或目录路径",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="干运行：解析并验证数据，不写入数据库",
    )


    args = parser.parse_args()

    # 查找文件
    input_path = args.input_path
    if os.path.isdir(input_path):
        files = find_bank_files(input_path)
        if not files:
            print(f"No bank files found in: {input_path}")
            sys.exit(1)
        print(f"Found {len(files)} files to process")
    else:
        files = [input_path]

    # 执行
    if args.dry_run:
        for f in files:
            dry_run_import(f)
    elif args.verify:
        for f in files:
            verify_import(f)
    else:
        for f in files:
            result = do_import(f)
            print(f"\n✓ Import completed: {result['row_count']} rows, source_file_id={result['source_file_id']}")


if __name__ == "__main__":
    main()
